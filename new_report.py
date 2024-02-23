import os
import pandas as pd
import mysql.connector
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Obtener la fecha actual
current_date = datetime.now().strftime('%d-%m-%Y')

# Definir el nombre de la carpeta con la fecha del día
folder_name = f"reporte_{current_date}"

# Obtener la ruta del directorio actual
current_directory = os.getcwd()

# Definir la ruta completa de la carpeta donde se guardarán los reportes
report_folder = os.path.join(current_directory, folder_name)

# Crear la carpeta si no existe
if not os.path.exists(report_folder):
    os.makedirs(report_folder)

# Especificar tus credenciales y detalles de conexión
config = {
    "host": "127.0.0.1",
    "port": "3307",
    "database": "hello_mysql",
    "user": "root",
    "password": "123456789" 
}

# Conectar a la base de datos
connection = mysql.connector.connect(**config)
cursor = connection.cursor()

# Define tu consulta SQL
query = "call p_all_users();"
cursor.execute(query)
result = cursor.fetchall()

# Convertir el resultado a un DataFrame de pandas
df = pd.DataFrame(result, columns=[i[0] for i in cursor.description])

# Guardar DataFrame en un archivo Excel dentro de la carpeta de reportes
excel_file = os.path.join(report_folder, f'reporte_{current_date}.xlsx')
df.to_excel(excel_file, index=False)

# Cerrar la conexión a la base de datos
cursor.close()
connection.close()

# Cargar el libro de trabajo de Excel
wb = load_workbook(excel_file)

# Obtener la hoja de cálculo activa
ws = wb.active

# Establecer el color de fondo de las cabeceras de las columnas en rojo
for cell in ws[1]:
    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.font = Font(color="FFFFFF")

# Guardar los cambios en el archivo Excel
wb.save(excel_file)

print(f"¡Reporte generado exitosamente y guardado en la carpeta '{folder_name}' como reporte.xlsx!")
